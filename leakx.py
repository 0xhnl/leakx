#!/usr/bin/env python3
import argparse
import csv
import hashlib
import re
import sys
import time
import random
import json
from urllib.parse import urlparse
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REQUEST_TIMEOUT = 20
API_BASE_URL = "https://api.leakradar.io"
DEFAULT_KEY_PATH = Path(__file__).resolve().parent / "keys.txt"
DEFAULT_LOOKUP = {
    "date_compromised": "Not Found",
    "computer_name": "Not Found",
    "operating_system": "Not Found",
    "malware_path": "Not Found",
}
VERBOSE = False


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge leak credential CSV files or LeakRadar API results into one styled XLSX file."
    )
    parser.add_argument(
        "-ff",
        "--folder",
        required=False,
        help="Folder containing CSV files. Defaults to ./files if it exists, else current directory.",
    )
    parser.add_argument(
        "-d",
        "--domain",
        required=False,
        help="Domain to filter CSVs by filename, or to fetch from LeakRadar API when -ff is omitted.",
    )
    parser.add_argument(
        "-dl",
        "--domain-list",
        required=False,
        help="Path to a text file with one domain per line (API mode or CSV filtering).",
    )
    parser.add_argument(
        "-i",
        "--ip",
        required=False,
        help="IP address to search via LeakRadar API (API mode only).",
    )
    parser.add_argument(
        "-il",
        "--ip-list",
        required=False,
        help="Path to a text file with one IP per line (API mode only).",
    )
    parser.add_argument(
        "-e",
        "--email",
        required=False,
        help="Email address to search via LeakRadar API (API mode only).",
    )
    parser.add_argument(
        "-el",
        "--email-list",
        required=False,
        help="Path to a text file with one email per line (API mode only).",
    )
    parser.add_argument(
        "-u",
        "--username",
        required=False,
        help="Username to search via LeakRadar API (API mode only).",
    )
    parser.add_argument(
        "-ul",
        "--username-list",
        required=False,
        help="Path to a text file with one username per line (API mode only).",
    )
    parser.add_argument(
        "-p",
        "--hash",
        "--password",
        required=False,
        dest="hash",
        help="Plaintext credential/password to search (auto-hashed to SHA-1 and checked via LeakRadar /password-range).",
    )
    parser.add_argument(
        "--hash-list",
        "--password-list",
        required=False,
        dest="hash_list",
        help="Path to a text file with one plaintext credential/password per line.",
    )
    parser.add_argument(
        "--category",
        required=False,
        choices=["employees", "customers", "third_parties", "all"],
        default="all",
        help="Domain search category filter: employees, customers, third_parties, or all (default: all).",
    )
    parser.add_argument(
        "--tenant",
        action="store_true",
        help="Expand domain(s) using tenant-domains and search all discovered domains.",
    )
    parser.add_argument(
        "--auto-unlock",
        action="store_true",
        help="API mode only: auto-unlock locked items on each page (consumes points).",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=50000,
        help="API mode only: maximum number of result rows to write (default: 50000).",
    )
    parser.add_argument(
        "--key",
        default=str(DEFAULT_KEY_PATH),
        help="API mode only: path to LeakRadar API key file (default: ./keys.txt).",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="Print progress while searching.",
    )
    parser.add_argument(
        "-o",
        "--output",
        required=True,
        help="Output XLSX file path.",
    )
    return parser.parse_args()


def read_csv_rows(file_path: Path) -> list[dict[str, str]]:
    encodings = ("utf-8-sig", "utf-8", "latin-1")
    last_error = None

    for enc in encodings:
        try:
            with file_path.open("r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    return []

                rows: list[dict[str, str]] = []
                for row in reader:
                    normalized = {k.strip(): (v or "").strip() for k, v in row.items() if k}
                    rows.append(normalized)
                return rows
        except UnicodeDecodeError as err:
            last_error = err

    raise RuntimeError(f"Cannot decode {file_path}: {last_error}")


def clean_string(value):
    if not isinstance(value, str):
        return "N/A"

    value = value.replace("\u2022", "REDACTED")
    while "REDACTEDREDACTED" in value:
        value = value.replace("REDACTEDREDACTED", "REDACTED")

    return value.strip()


def is_email(value: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", value.strip()))


def lookup_username(username: str) -> dict[str, str]:
    url = f"https://cavalier.hudsonrock.com/api/json/v2/osint-tools/search-by-username?username={username}"
    try:
        response = requests.get(url, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        data = response.json()
        return {
            "date_compromised": data.get("date_compromised", "N/A"),
            "computer_name": data.get("computer_name", "N/A"),
            "operating_system": data.get("operating_system", "N/A"),
            "malware_path": data.get("malware_path", "N/A"),
        }
    except Exception:
        return DEFAULT_LOOKUP.copy()


def lookup_email(email: str) -> dict[str, str]:
    url = f"https://cavalier.hudsonrock.com/api/json/v2/osint-tools/search-by-email?email={email}"
    try:
        response = requests.get(url, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        data = response.json()
        if "stealers" in data and isinstance(data["stealers"], list) and data["stealers"]:
            stealer = data["stealers"][0]
            if isinstance(stealer, dict):
                return {
                    "date_compromised": stealer.get("date_compromised", "N/A"),
                    "computer_name": clean_string(stealer.get("computer_name", "N/A")),
                    "operating_system": clean_string(stealer.get("operating_system", "N/A")),
                    "malware_path": clean_string(stealer.get("malware_path", "N/A")),
                }
    except Exception:
        pass

    return DEFAULT_LOOKUP.copy()


def enrich_username_field(
    username_value: str,
    cache: dict[str, dict[str, str]],
    is_email_hint: bool | None = None,
) -> dict[str, str]:
    key = (username_value or "").strip()
    if not key:
        return DEFAULT_LOOKUP.copy()

    if key in cache:
        return cache[key]

    if is_email_hint is True or (is_email_hint is None and is_email(key)):
        result = lookup_email(key)
    else:
        result = lookup_username(key)

    cache[key] = result
    return result


def read_api_key(path: Path) -> str:
    if not path.exists() and path.name == "key.txt":
        alt = path.with_name("keys.txt")
        if alt.exists():
            path = alt
    if not path.exists():
        raise FileNotFoundError(f"API key file not found: {path}")
    key = path.read_text(encoding="utf-8").strip()
    if not key:
        raise RuntimeError(f"API key file is empty: {path}")
    return key


def fetch_domain_leaks(domain: str, api_key: str, auto_unlock: bool, category: str = "all") -> tuple[list[dict], int]:
    headers = {"Authorization": f"Bearer {api_key}"}
    page = 1
    page_size = 1000
    all_items: list[dict] = []
    points_consumed = 0

    while True:
        params = {"page": page, "page_size": page_size}
        if auto_unlock:
            params["auto_unlock"] = "true"

        url = f"{API_BASE_URL}/search/domain/{domain}/{category}"
        if VERBOSE:
            print(f"[domain] {domain} page={page} page_size={page_size}", file=sys.stderr)
        response = requests.get(url, headers=headers, params=params, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        data = response.json()

        items = data.get("items", [])
        all_items.extend(items)
        if VERBOSE:
            total = data.get("total")
            print(f"[domain] {domain} got {len(items)} items (total={total})", file=sys.stderr)

        points = data.get("auto_unlock_points_consumed")
        if isinstance(points, int):
            points_consumed += points

        total = data.get("total")
        if total is not None and len(all_items) >= total:
            break
        if not items or len(items) < page_size:
            break

        page += 1

    return all_items, points_consumed


def fetch_ip_leaks(ip: str, api_key: str, auto_unlock: bool) -> tuple[list[dict], int]:
    headers = {"Authorization": f"Bearer {api_key}"}
    page = 1
    page_size = 1000
    all_items: list[dict] = []
    points_consumed = 0

    if auto_unlock:
        # Walk all pages to trigger auto-unlock on each page.
        while True:
            params = {"page": page, "page_size": page_size, "auto_unlock": True}
            url = f"{API_BASE_URL}/search/advanced"
            payload = {"url_host": [ip]}
            if VERBOSE:
                print(
                    f"[ip] {ip} page={page} page_size={page_size} filter=url_host (auto_unlock)",
                    file=sys.stderr,
                )
            response = requests.post(
                url, headers=headers, params=params, json=payload, timeout=REQUEST_TIMEOUT
            )
            response.raise_for_status()
            data = response.json()

            points = data.get("auto_unlock_points_consumed")
            if isinstance(points, int):
                points_consumed += points

            items = data.get("items", [])
            total = data.get("total")
            if total is not None and (page * page_size) >= total:
                break
            if not items or len(items) < page_size:
                break
            page += 1

        # After auto-unlock, fetch only unlocked items using the unlocked endpoint.
        unlocked_items = fetch_unlocked_advanced({"url_host": [ip]}, api_key)
        return unlocked_items, points_consumed

    while True:
        params = {"page": page, "page_size": page_size}
        if auto_unlock:
            params["auto_unlock"] = "true"

        url = f"{API_BASE_URL}/search/advanced"
        payload = {"url_host": [ip]}
        if VERBOSE:
            print(f"[ip] {ip} page={page} page_size={page_size} filter=url_host", file=sys.stderr)
        response = requests.post(
            url, headers=headers, params=params, json=payload, timeout=REQUEST_TIMEOUT
        )
        response.raise_for_status()
        data = response.json()

        items = data.get("items", [])
        all_items.extend(items)
        if VERBOSE:
            total = data.get("total")
            print(f"[ip] {ip} got {len(items)} items (total={total})", file=sys.stderr)

        points = data.get("auto_unlock_points_consumed")
        if isinstance(points, int):
            points_consumed += points

        total = data.get("total")
        if total is not None and len(all_items) >= total:
            break
        if not items or len(items) < page_size:
            break

        page += 1

    return all_items, points_consumed


def fetch_unlocked_advanced(filters: dict, api_key: str) -> list[dict]:
    headers = {"Authorization": f"Bearer {api_key}"}
    page = 1
    page_size = 1000
    all_items: list[dict] = []

    while True:
        params = {"page": page, "page_size": page_size}
        for key, value in (filters or {}).items():
            params[key] = value

        url = f"{API_BASE_URL}/profile/unlocked/advanced"
        if VERBOSE:
            print(f"[advanced-unlocked] page={page} page_size={page_size}", file=sys.stderr)
        response = requests.get(url, headers=headers, params=params, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        data = response.json()

        items = data.get("items", [])
        all_items.extend(items)

        total = data.get("total")
        if total is not None and len(all_items) >= total:
            break
        if not items or len(items) < page_size:
            break

        page += 1

    return all_items


def fetch_password_hashes(password: str, api_key: str) -> list[dict]:
    """Hash password to SHA-1, call /password-range, return matching hash rows."""
    sha1 = hashlib.sha1(password.encode("utf-8")).hexdigest().upper()
    prefix = sha1[:5]
    suffix = sha1[5:]
    headers = {"Authorization": f"Bearer {api_key}"}
    url = f"{API_BASE_URL}/password-range"
    params = {"prefix": prefix, "suffix_only": "false"}
    if VERBOSE:
        print(f"[password] '{password}' -> SHA1={sha1} prefix={prefix}", file=sys.stderr)
    response = requests.get(url, headers=headers, params=params, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    data = response.json()
    hashes = data.get("hashes", [])
    # Find the exact match for this password's full hash
    results = []
    for entry in hashes:
        h = entry.get("hash", "").upper()
        if h == sha1:
            results.append({
                "PASSWORD": password,
                "SHA1_HASH": sha1,
                "COUNT": entry.get("count", 0),
            })
            break
    if not results:
        results.append({
            "PASSWORD": password,
            "SHA1_HASH": sha1,
            "COUNT": 0,
        })
    return results

def fetch_password_leaks(
    password: str, api_key: str, auto_unlock: bool, limit: int | None
) -> tuple[list[dict], int]:
    """Search by plaintext password via /search/advanced, returning leak items."""
    headers = {"Authorization": f"Bearer {api_key}"}
    page = 1
    page_size = 1000
    all_items: list[dict] = []
    points_consumed = 0

    while True:
        params = {"page": page, "page_size": page_size}
        if auto_unlock:
            params["auto_unlock"] = "true"

        url = f"{API_BASE_URL}/search/advanced"
        payload = {"password": [password]}
        if VERBOSE:
            print(f"[password] page={page} page_size={page_size} filter=password", file=sys.stderr)
        response = requests.post(
            url, headers=headers, params=params, json=payload, timeout=REQUEST_TIMEOUT
        )
        response.raise_for_status()
        data = response.json()

        items = data.get("items", [])
        all_items.extend(items)

        points = data.get("auto_unlock_points_consumed")
        if isinstance(points, int):
            points_consumed += points

        total = data.get("total")
        if limit is not None and len(all_items) >= limit:
            all_items = all_items[:limit]
            break
        if total is not None and len(all_items) >= total:
            break
        if not items or len(items) < page_size:
            break

        page += 1

    return all_items, points_consumed

def fetch_email_leaks(email: str, api_key: str, auto_unlock: bool) -> tuple[list[dict], int]:
    headers = {"Authorization": f"Bearer {api_key}"}
    page = 1
    page_size = 100  # /search/email max page_size is 100
    all_items: list[dict] = []
    points_consumed = 0

    while True:
        params = {"page": page, "page_size": page_size}
        if auto_unlock:
            params["auto_unlock"] = "true"

        url = f"{API_BASE_URL}/search/email"
        payload = {"email": email}
        if VERBOSE:
            print(f"[email] {email} page={page} page_size={page_size}", file=sys.stderr)
        response = requests.post(url, headers=headers, params=params, json=payload, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        data = response.json()

        items = data.get("items", [])
        all_items.extend(items)
        if VERBOSE:
            total = data.get("total")
            print(f"[email] {email} got {len(items)} items (total={total})", file=sys.stderr)

        points = data.get("auto_unlock_points_consumed")
        if isinstance(points, int):
            points_consumed += points

        total = data.get("total")
        if total is not None and len(all_items) >= total:
            break
        if not items or len(items) < page_size:
            break

        page += 1

    return all_items, points_consumed


def autosize_columns(ws, headers: list[str]) -> None:
    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            val = row[0].value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))

        width = min(max(max_len + 2, 12), 70)
        ws.column_dimensions[get_column_letter(idx)].width = width


def style_sheet(ws, headers: list[str], row_count: int) -> None:
    title_font = Font(name="Titillium Web", size=15, bold=True, color="0B132B")
    header_font = Font(name="Titillium Web", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Titillium Web", size=10, color="111111")

    header_fill = PatternFill(fill_type="solid", fgColor="1D4ED8")
    alt_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")

    thin_side = Side(style="thin", color="D1D5DB")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx in range(2, row_count + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(row_count, 1)}"
    ws.row_dimensions[1].height = 22

    autosize_columns(ws, headers)


def read_domain_list(path: Path) -> list[str]:
    domains: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        value = line.strip()
        if value and not value.startswith("#"):
            domains.append(value)
    return domains


def read_ip_list(path: Path) -> list[str]:
    ips: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        value = line.strip()
        if value and not value.startswith("#"):
            ips.append(value)
    return ips


def read_email_list(path: Path) -> list[str]:
    emails: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        value = line.strip()
        if value and not value.startswith("#"):
            emails.append(value)
    return emails


def expand_tenant_domains(domains: list[str]) -> list[str]:
    expanded: set[str] = set()
    for domain in domains:
        domain = domain.strip().lower()
        expanded.add(domain)
        try:
            if VERBOSE:
                print(f"[tenant] resolving {domain}", file=sys.stderr)
            oidc_url = f"https://login.microsoftonline.com/{domain}/.well-known/openid-configuration"
            resp = requests.get(
                oidc_url,
                timeout=REQUEST_TIMEOUT,
                headers={
                    "user-agent": (
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/139.0.0.0 Safari/537.36"
                    )
                },
            )
            if resp.status_code != 200:
                if VERBOSE:
                    print(f"[tenant] {domain} oidc status={resp.status_code}", file=sys.stderr)
                continue
            try:
                data = resp.json()
            except json.JSONDecodeError:
                if VERBOSE:
                    print(f"[tenant] {domain} oidc response not json", file=sys.stderr)
                continue

            token_endpoint = data.get("token_endpoint", "")
            issuer = data.get("issuer", "")
            auth_endpoint = data.get("authorization_endpoint", "")
            tenant_id = ""
            for value in (token_endpoint, issuer, auth_endpoint):
                if not value:
                    continue
                parts = value.split("/")
                if len(parts) > 3:
                    tenant_id = parts[3]
                    break
            if not tenant_id:
                if VERBOSE:
                    print(f"[tenant] {domain} no tenant_id found", file=sys.stderr)
                continue

            time.sleep(1 + random.random() * 10)
            tenant_url = f"https://tenant-api.micahvandeusen.com/search?tenant_id={tenant_id}"
            t_resp = requests.get(
                tenant_url,
                timeout=REQUEST_TIMEOUT,
                headers={
                    "user-agent": (
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/139.0.0.0 Safari/537.36"
                    )
                },
            )
            if t_resp.status_code != 200:
                if VERBOSE:
                    print(f"[tenant] {domain} tenant-api status={t_resp.status_code}", file=sys.stderr)
                continue
            t_data = t_resp.json()
            found = 0
            for value in t_data.get("domains", []):
                if isinstance(value, str) and value.strip():
                    expanded.add(value.strip().lower())
                    found += 1
            if VERBOSE:
                print(f"[tenant] {domain} found {found} domains", file=sys.stderr)
        except Exception:
            if VERBOSE:
                print(f"[tenant] {domain} error during expansion", file=sys.stderr)
            continue

    return sorted(expanded)


def build_report_from_csv(
    folder: Path, domain_filters: list[str] | None
) -> tuple[list[dict[str, str]], list[str]]:
    csv_files = sorted(folder.glob("*.csv"))
    if not csv_files:
        raise FileNotFoundError(f"No CSV files found in: {folder}")

    if domain_filters:
        csv_files = [
            csv_file
            for csv_file in csv_files
            if len(csv_file.name.split("_")) > 1
            and csv_file.name.split("_")[1] in domain_filters
        ]
        if not csv_files:
            raise FileNotFoundError(
                f"No CSV files for domains '{', '.join(domain_filters)}' found in: {folder}"
            )

    all_rows: list[dict[str, str]] = []
    all_columns: list[str] = []
    cache: dict[str, dict[str, str]] = {}

    for csv_file in csv_files:
        parts = csv_file.name.split("_")
        domain = parts[1] if len(parts) > 1 else ""
        category = parts[2] if len(parts) > 2 else ""

        rows = read_csv_rows(csv_file)
        for row in rows:
            for key in row.keys():
                if key not in all_columns:
                    all_columns.append(key)

            row_data = {
                "DOMAIN": domain,
                "CATEGORY": category,
            }
            row_data.update(row)

            username_value = row.get("USERNAME", "")
            lookup = enrich_username_field(username_value, cache)
            row_data.update(lookup)
            for key in lookup.keys():
                if key not in all_columns:
                    all_columns.append(key)
            all_rows.append(row_data)

    return all_rows, all_columns


def build_report_from_api(
    domains: list[str], api_key: str, auto_unlock: bool, category: str = "all"
) -> tuple[list[dict[str, str]], list[str], int]:
    all_rows: list[dict[str, str]] = []
    all_columns: list[str] = []
    cache: dict[str, dict[str, str]] = {}
    total_points = 0

    for domain in domains:
        if VERBOSE:
            print(f"[domain] start {domain}", file=sys.stderr)
        items, points_consumed = fetch_domain_leaks(domain, api_key, auto_unlock, category)
        total_points += points_consumed
        for item in items:
            username_value = item.get("username") or item.get("username_masked") or ""
            row_data = {
                "DOMAIN": domain,
                "CATEGORY": item.get("category", ""),
                "URL": item.get("url", ""),
                "USERNAME": username_value,
                "PASSWORD": item.get("password", ""),
                "ADDED_AT": item.get("added_at", ""),
            }

            for key in row_data.keys():
                if key not in all_columns:
                    all_columns.append(key)

            is_email_hint = item.get("is_email")
            lookup = enrich_username_field(username_value, cache, is_email_hint=is_email_hint)
            row_data.update(lookup)
            for key in lookup.keys():
                if key not in all_columns:
                    all_columns.append(key)
            all_rows.append(row_data)

    return all_rows, all_columns, total_points


def build_report_from_ip(
    ips: list[str], api_key: str, auto_unlock: bool
) -> tuple[list[dict[str, str]], list[str], int]:
    all_rows: list[dict[str, str]] = []
    all_columns: list[str] = []
    cache: dict[str, dict[str, str]] = {}
    total_points = 0

    for ip in ips:
        if VERBOSE:
            print(f"[ip] start {ip}", file=sys.stderr)
        items, points_consumed = fetch_ip_leaks(ip, api_key, auto_unlock)
        total_points += points_consumed
        for item in items:
            username_value = item.get("username") or item.get("username_masked") or ""
            row_data = {
                "DOMAIN": "",
                "CATEGORY": "",
                "IP": ip,
                "URL": item.get("url", ""),
                "USERNAME": username_value,
                "PASSWORD": item.get("password", ""),
                "ADDED_AT": item.get("added_at", ""),
            }

            for key in row_data.keys():
                if key not in all_columns:
                    all_columns.append(key)

            is_email_hint = item.get("is_email")
            lookup = enrich_username_field(username_value, cache, is_email_hint=is_email_hint)
            row_data.update(lookup)
            for key in lookup.keys():
                if key not in all_columns:
                    all_columns.append(key)
            all_rows.append(row_data)

    return all_rows, all_columns, total_points


def build_report_from_email(
    emails: list[str], api_key: str, auto_unlock: bool
) -> tuple[list[dict[str, str]], list[str], int]:
    all_rows: list[dict[str, str]] = []
    all_columns: list[str] = []
    cache: dict[str, dict[str, str]] = {}
    total_points = 0

    for email in emails:
        if VERBOSE:
            print(f"[email] start {email}", file=sys.stderr)
        items, points_consumed = fetch_email_leaks(email, api_key, auto_unlock)
        total_points += points_consumed
        for item in items:
            username_value = item.get("username") or item.get("username_masked") or email
            row_data = {
                "DOMAIN": "",
                "CATEGORY": item.get("category", ""),
                "EMAIL": email,
                "URL": item.get("url", ""),
                "USERNAME": username_value,
                "PASSWORD": item.get("password", ""),
                "ADDED_AT": item.get("added_at", ""),
            }

            for key in row_data.keys():
                if key not in all_columns:
                    all_columns.append(key)

            is_email_hint = item.get("is_email", True)
            lookup = enrich_username_field(username_value, cache, is_email_hint=is_email_hint)
            row_data.update(lookup)
            for key in lookup.keys():
                if key not in all_columns:
                    all_columns.append(key)
            all_rows.append(row_data)

    return all_rows, all_columns, total_points


def build_report_from_username(
    usernames: list[str], api_key: str, auto_unlock: bool
) -> tuple[list[dict[str, str]], list[str], int]:
    """Username lookup — reuses fetch_email_leaks (same POST /search/email endpoint)."""
    all_rows: list[dict[str, str]] = []
    all_columns: list[str] = []
    cache: dict[str, dict[str, str]] = {}
    total_points = 0

    for username in usernames:
        if VERBOSE:
            print(f"[username] start {username}", file=sys.stderr)
        items, points_consumed = fetch_email_leaks(username, api_key, auto_unlock)
        total_points += points_consumed
        for item in items:
            username_value = item.get("username") or item.get("username_masked") or username
            row_data = {
                "DOMAIN": "",
                "CATEGORY": item.get("category", ""),
                "SEARCHED_USERNAME": username,
                "URL": item.get("url", ""),
                "USERNAME": username_value,
                "PASSWORD": item.get("password", ""),
                "ADDED_AT": item.get("added_at", ""),
            }

            for key in row_data.keys():
                if key not in all_columns:
                    all_columns.append(key)

            is_email_hint = item.get("is_email", False)
            lookup = enrich_username_field(username_value, cache, is_email_hint=is_email_hint)
            row_data.update(lookup)
            for key in lookup.keys():
                if key not in all_columns:
                    all_columns.append(key)
            all_rows.append(row_data)

    return all_rows, all_columns, total_points


def build_report_from_password(
    passwords: list[str], api_key: str, auto_unlock: bool, limit: int | None
) -> tuple[list[dict[str, str]], list[str], int]:
    """For each plaintext password, compute SHA1+range count and fetch affected leak items."""
    all_rows: list[dict[str, str]] = []
    columns: list[str] = []
    total_points = 0

    for password in passwords:
        pwd = (password or "").strip()
        if not pwd:
            continue

        sha1 = hashlib.sha1(pwd.encode("utf-8")).hexdigest().upper()
        range_rows = fetch_password_hashes(pwd, api_key)
        count = range_rows[0].get("COUNT", 0) if range_rows else 0

        items, points_consumed = fetch_password_leaks(pwd, api_key, auto_unlock, limit)
        total_points += points_consumed

        if not items:
            row = {
                "DOMAIN": "",
                "CATEGORY": "",
                "URL": "",
                "USERNAME": "",
                "PASSWORD": "",
                "ADDED_AT": "",
                "SEARCHED_PASSWORD": pwd,
                "SHA1_HASH": sha1,
                "COUNT": count,
            }
            for key in row.keys():
                if key not in columns:
                    columns.append(key)
            all_rows.append(row)
            continue

        for item in items:
            url_value = item.get("url", "") or ""
            domain_value = ""
            try:
                parsed = urlparse(url_value)
                domain_value = parsed.hostname or ""
            except Exception:
                domain_value = ""

            username_value = item.get("username") or item.get("username_masked") or ""
            row = {
                "DOMAIN": domain_value,
                "CATEGORY": item.get("category", ""),
                "URL": url_value,
                "USERNAME": username_value,
                "PASSWORD": item.get("password", ""),
                "ADDED_AT": item.get("added_at", ""),
                "SEARCHED_PASSWORD": pwd,
                "SHA1_HASH": sha1,
                "COUNT": count,
            }
            for key in row.keys():
                if key not in columns:
                    columns.append(key)
            all_rows.append(row)
        if limit is not None and len(all_rows) >= limit:
            all_rows = all_rows[:limit]
            break

    return all_rows, columns, total_points

def build_headers(columns: list[str]) -> list[str]:
    ordered_headers = [
        "DOMAIN",
        "CATEGORY",
        "URL",
        "USERNAME",
        "PASSWORD",
        "ADDED_AT",
        "date_compromised",
        "computer_name",
        "operating_system",
        "malware_path",
    ]
    if "IP" in columns:
        ordered_headers.insert(2, "IP")
    if "EMAIL" in columns:
        ordered_headers.insert(2, "EMAIL")
    headers = [h for h in ordered_headers if h in columns]
    extra_columns = [c for c in columns if c not in headers]
    headers.extend(extra_columns)
    return headers


def write_sheet(ws, headers: list[str], rows: list[dict[str, str]]) -> None:
    row_pointer = 2
    for row in rows:
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_pointer, column=col_idx, value=row.get(header, ""))
        row_pointer += 1
    style_sheet(ws, headers, row_pointer - 1)


def write_report(output: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    headers = build_headers(columns)
    wb = Workbook()
    ws = wb.active
    ws.title = "Leak Report"
    write_sheet(ws, headers, rows)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def write_report_by_domain(output: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    headers = build_headers(columns)
    wb = Workbook()
    first = True
    for domain in sorted({r.get("DOMAIN", "") for r in rows}):
        domain_rows = [r for r in rows if r.get("DOMAIN", "") == domain]
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        ws.title = domain[:31] or "unknown"
        write_sheet(ws, headers, domain_rows)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    args = parse_args()
    global VERBOSE
    VERBOSE = bool(args.verbose)

    output = Path(args.output).expanduser().resolve()

    try:
        folder: Path | None = None
        domain_filters: list[str] | None = None
        if args.domain_list:
            domain_filters = read_domain_list(Path(args.domain_list).expanduser().resolve())
        elif args.domain:
            domain_filters = [args.domain.strip()]

        if domain_filters and args.tenant:
            domain_filters = expand_tenant_domains(domain_filters)
            print(
                f"[tenant] total domains after expansion: {len(domain_filters)}",
                file=sys.stderr,
            )

        ip_filters: list[str] | None = None
        if args.ip_list:
            ip_filters = read_ip_list(Path(args.ip_list).expanduser().resolve())
        elif args.ip:
            ip_filters = [args.ip.strip()]

        email_filters: list[str] | None = None
        if args.email_list:
            email_filters = read_email_list(Path(args.email_list).expanduser().resolve())
        elif args.email:
            email_filters = [args.email.strip()]

        username_filters: list[str] | None = None
        if args.username_list:
            username_filters = read_email_list(Path(args.username_list).expanduser().resolve())
        elif args.username:
            username_filters = [args.username.strip()]

        hash_list: list[str] | None = None
        if args.hash_list:
            hash_list = read_email_list(Path(args.hash_list).expanduser().resolve())
        elif args.hash:
            hash_list = [args.hash]

        if args.folder:
            folder = Path(args.folder).expanduser().resolve()
        elif not domain_filters and not ip_filters and not email_filters and not username_filters and not hash_list:
            default_folder = Path.cwd() / "files"
            folder = default_folder if default_folder.exists() else Path.cwd()

        if folder is not None:
            if ip_filters:
                print("Error: IP search is API-only. Omit -ff for IP mode.", file=sys.stderr)
                return 1
            if email_filters:
                print("Error: Email search is API-only. Omit -ff for email mode.", file=sys.stderr)
                return 1
            if not folder.exists() or not folder.is_dir():
                print(f"Error: Folder not found or not a directory: {folder}", file=sys.stderr)
                return 1

            rows, columns = build_report_from_csv(folder, domain_filters)
            unique_domains = {r.get("DOMAIN", "") for r in rows}
            if len(unique_domains) > 1:
                write_report_by_domain(output, rows, columns)
            else:
                write_report(output, rows, columns)

            if domain_filters:
                print(
                    f"Done: merged {len(rows)} rows from CSVs for domains '{', '.join(domain_filters)}'."
                )
            else:
                print(f"Done: merged {len(rows)} rows from CSVs.")
            return 0

        active_modes = sum([
            bool(domain_filters),
            bool(ip_filters),
            bool(email_filters),
            bool(username_filters),
            bool(hash_list),
        ])
        if active_modes == 0:
            print(
                "Error: Provide -ff for CSV mode or -d/-dl/-i/-il/-e/-el/-u/-ul/-p/--hash-list for API mode.",
                file=sys.stderr,
            )
            return 1
        if active_modes > 1:
            print(
                "Error: Choose only one search mode at a time.",
                file=sys.stderr,
            )
            return 1

        api_key = read_api_key(Path(args.key).expanduser().resolve())
        if domain_filters:
            rows, columns, points_consumed = build_report_from_api(
                domain_filters, api_key, args.auto_unlock, args.category
            )
            summary = f"{len(domain_filters)} domain(s) [{args.category}]"
            if len(domain_filters) > 1:
                write_report_by_domain(output, rows, columns)
            else:
                write_report(output, rows, columns)
        elif ip_filters:
            rows, columns, points_consumed = build_report_from_ip(
                ip_filters, api_key, args.auto_unlock
            )
            summary = f"{len(ip_filters)} IP(s)"
            write_report(output, rows, columns)
        elif email_filters:
            rows, columns, points_consumed = build_report_from_email(
                email_filters, api_key, args.auto_unlock
            )
            summary = f"{len(email_filters)} email(s)"
            write_report(output, rows, columns)
        elif username_filters:
            rows, columns, points_consumed = build_report_from_username(
                username_filters, api_key, args.auto_unlock
            )
            summary = f"{len(username_filters)} username(s)"
            write_report(output, rows, columns)
        else:
            # hash mode
            rows, columns, points_consumed = build_report_from_password(
                hash_list or [], api_key, args.auto_unlock, args.limit
            )
            summary = f"{len(hash_list or [])} hash lookup(s)"
            write_report(output, rows, columns)
            print(f"Done: checked {len(rows)} credential/password hash(es). Results in {output}")
            return 0

        if args.auto_unlock:
            note = f"Auto-unlock points consumed: {points_consumed}"
        else:
            note = "Auto-unlock points consumed: 0 (auto-unlock disabled)"
        print(f"Done: fetched {len(rows)} rows for {summary}. {note}")
        return 0
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
